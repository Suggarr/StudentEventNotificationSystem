from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QTableWidgetItem, QHeaderView,  QListWidgetItem, QFileDialog, QMessageBox, QAbstractItemView
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import Qt
import smtplib, socket, re
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os


class Ui_MainForm(object):
    def __init__(self):
        self.file_paths = []
    def setupUi(self, MainForm): # Создание окна
        MainForm.setObjectName("MainForm")
        MainForm.resize(1151, 939)
        MainForm.setMinimumSize(QtCore.QSize(1151, 939))
        MainForm.setMaximumSize(QtCore.QSize(1151, 939))
        MainForm.setMouseTracking(False)
        MainForm.setTabletTracking(False)
        MainForm.setFocusPolicy(QtCore.Qt.ClickFocus)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("resources/mail_ikonka2.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainForm.setWindowIcon(icon)
        MainForm.setStyleSheet("*{\n"
"    color: #FFF;\n"
"    background-color: rgb(68, 136, 204);\n"
"}\n"
"QMenu\n"
"{\n"
"    background-color: rgb(0, 58, 176);\n"
"}\n"
"QMenuBar\n"
"{\n"
"    background-color: rgb(0, 58, 176);\n"
"}\n"
"QMenuBar::item:selected {\n"
"    background-color: rgba(55, 111, 166); /* затемнение с помощью полупрозрачного цвета */\n"
"    color: white; /* цвет текста */\n"
"}\n"
"QMenu::item:selected {\n"
"    background-color: rgba(55, 111, 166); /* затемнение с помощью полупрозрачного цвета */\n"
"    color: white; /* цвет текста */\n"
"}\n"
"QPushButton\n"
"{\n"
"    font: 14pt \"MS Shell Dlg 2\";\n"
"    color: rgb(255, 255, 255);\n"
"    background-color: rgb(0, 58, 176);\n"
"    border-radius: 10px;                   \n"
"    border: 2px solid #094065;\n"
"}\n"
"QPushButton::pressed\n"
"{\n"
"    font: 14pt \"MS Shell Dlg 2\";\n"
"    color: rgb(255, 0, 0);\n"
"    background-color: rgb(0, 158, 176);\n"
"    border-radius: 10px;\n"
"    border: 2px solid #FF0000;\n"
"}\n"
"QPushButton::released\n"
"{\n"
"    font: 14pt \"MS Shell Dlg 2\";\n"
"    color: rgb(255, 255, 255);\n"
"    background-color: rgb(0, 58, 176);\n"
"    border-radius: 10px;\n"
"    border: 2px solid #094065;\n"
"}\n"
"\n"
"")
        self.centralwidget = QtWidgets.QWidget(MainForm)
        self.centralwidget.setObjectName("centralwidget")
        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(10, 10, 701, 821))
        self.groupBox.setStyleSheet("")
        self.groupBox.setObjectName("groupBox")
        self.label_2 = QtWidgets.QLabel(self.groupBox)
        self.label_2.setGeometry(QtCore.QRect(20, 30, 81, 21))
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.groupBox)
        self.label_3.setGeometry(QtCore.QRect(200, 30, 41, 21))
        self.label_3.setObjectName("label_3")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.groupBox)
        self.lineEdit_2.setGeometry(QtCore.QRect(20, 60, 141, 31))
        self.lineEdit_2.setStyleSheet("background-color: rgb(217, 217, 217);\n"
"color: rgb(0, 0, 0);\n"
"border-radius: 10px; ")
        self.lineEdit_2.setText("")
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.lineEdit_3 = QtWidgets.QLineEdit(self.groupBox)
        self.lineEdit_3.setGeometry(QtCore.QRect(200, 60, 141, 31))
        self.lineEdit_3.setStyleSheet("background-color: rgb(217, 217, 217);\n"
"color: rgb(0, 0, 0);\n"
"border-radius: 10px; ")
        self.lineEdit_3.setText("")
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.label_4 = QtWidgets.QLabel(self.groupBox)
        self.label_4.setGeometry(QtCore.QRect(360, 30, 81, 21))
        self.label_4.setObjectName("label_4")
        self.lineEdit_4 = QtWidgets.QLineEdit(self.groupBox)
        self.lineEdit_4.setGeometry(QtCore.QRect(360, 60, 141, 31))
        self.lineEdit_4.setStyleSheet("background-color: rgb(217, 217, 217);\n"
"color: rgb(0, 0, 0);\n"
"border-radius: 10px; ")
        self.lineEdit_4.setText("")
        self.lineEdit_4.setObjectName("lineEdit_4")
        self.lineEdit_5 = QtWidgets.QLineEdit(self.groupBox)
        self.lineEdit_5.setGeometry(QtCore.QRect(540, 60, 141, 31))
        self.lineEdit_5.setStyleSheet("background-color: rgb(217, 217, 217);\n"
"color: rgb(0, 0, 0);\n"
"border-radius: 10px; ")
        self.lineEdit_5.setText("")
        self.lineEdit_5.setObjectName("lineEdit_5")
        self.label_5 = QtWidgets.QLabel(self.groupBox)
        self.label_5.setGeometry(QtCore.QRect(540, 30, 61, 21))
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.groupBox)
        self.label_6.setGeometry(QtCore.QRect(20, 120, 51, 21))
        self.label_6.setObjectName("label_6")
        self.lineEdit_6 = QtWidgets.QLineEdit(self.groupBox)
        self.lineEdit_6.setGeometry(QtCore.QRect(20, 150, 141, 31))
        self.lineEdit_6.setStyleSheet("background-color: rgb(217, 217, 217);\n"
"color: rgb(0, 0, 0);\n"
"border-radius: 10px; ")
        self.lineEdit_6.setText("")
        self.lineEdit_6.setObjectName("lineEdit_6")
        self.add = QtWidgets.QPushButton(self.groupBox)
        self.add.setGeometry(QtCore.QRect(240, 140, 181, 51))
        self.add.setStyleSheet("")
        self.add.setCheckable(False)
        self.add.setObjectName("add")
        self.del_text = QtWidgets.QPushButton(self.groupBox)
        self.del_text.setGeometry(QtCore.QRect(500, 140, 181, 51))
        self.del_text.setStyleSheet("")
        self.del_text.setCheckable(False)
        self.del_text.setObjectName("del_text")
        self.tableWidget = QtWidgets.QTableWidget(self.groupBox)
        self.tableWidget.setGeometry(QtCore.QRect(20, 230, 661, 571))
        self.tableWidget.setStyleSheet("background-color: rgb(217, 217, 217);\n"
                                       "")
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.groupBox_2 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_2.setGeometry(QtCore.QRect(720, 10, 421, 201))
        self.groupBox_2.setObjectName("groupBox_2")
        self.label = QtWidgets.QLabel(self.groupBox_2)
        self.label.setGeometry(QtCore.QRect(10, 30, 391, 21))
        self.label.setObjectName("label")
        self.del_id = QtWidgets.QPushButton(self.groupBox_2)
        self.del_id.setGeometry(QtCore.QRect(120, 120, 181, 51))
        self.del_id.setStyleSheet("")
        self.del_id.setCheckable(False)
        self.del_id.setObjectName("del_id")
        self.lineEdit = QtWidgets.QLineEdit(self.groupBox_2)
        self.lineEdit.setGeometry(QtCore.QRect(20, 70, 381, 31))
        self.lineEdit.setStyleSheet("background-color: rgb(217, 217, 217);\n"
"color: rgb(0, 0, 0);\n"
"border-radius: 10px; ")
        self.lineEdit.setText("")
        self.lineEdit.setObjectName("lineEdit")
        self.groupBox_3 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_3.setGeometry(QtCore.QRect(720, 210, 421, 621))
        self.groupBox_3.setObjectName("groupBox_3")
        self.label_7 = QtWidgets.QLabel(self.groupBox_3)
        self.label_7.setGeometry(QtCore.QRect(50, 90, 331, 21))
        self.label_7.setObjectName("label_7")
        self.lineEdit_7 = QtWidgets.QLineEdit(self.groupBox_3)
        self.lineEdit_7.setGeometry(QtCore.QRect(20, 120, 381, 31))
        self.lineEdit_7.setStyleSheet("background-color: rgb(217, 217, 217);\n"
"color: rgb(0, 0, 0);\n"
"border-radius: 10px; ")
        self.lineEdit_7.setText("")
        self.lineEdit_7.setObjectName("lineEdit_7")
        self.label_8 = QtWidgets.QLabel(self.groupBox_3)
        self.label_8.setGeometry(QtCore.QRect(50, 160, 311, 21))
        self.label_8.setObjectName("label_8")
        self.label_9 = QtWidgets.QLabel(self.groupBox_3)
        self.label_9.setGeometry(QtCore.QRect(20, 20, 381, 21))
        self.label_9.setObjectName("label_9")
        self.lineEdit_9 = QtWidgets.QLineEdit(self.groupBox_3)
        self.lineEdit_9.setGeometry(QtCore.QRect(20, 50, 381, 31))
        self.lineEdit_9.setStyleSheet("background-color: rgb(217, 217, 217);\n"
"color: rgb(0, 0, 0);\n"
"border-radius: 10px; ")
        self.lineEdit_9.setText("")
        self.lineEdit_9.setObjectName("lineEdit_9")
        self.plainTextEdit = QtWidgets.QPlainTextEdit(self.groupBox_3)
        self.plainTextEdit.setGeometry(QtCore.QRect(20, 190, 381, 231))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.plainTextEdit.setFont(font)
        self.plainTextEdit.setStyleSheet("background-color: rgb(217, 217, 217);\n"
                                         "color: rgb(0, 0, 0);\n"
                                         "border-radius: 10px; ")
        self.plainTextEdit.setObjectName("plainTextEdit")
        self.attach = QtWidgets.QPushButton(self.groupBox_3)
        self.attach.setGeometry(QtCore.QRect(20, 490, 181, 51))
        self.attach.setStyleSheet("")
        self.attach.setCheckable(False)
        self.attach.setObjectName("attach")
        self.listWidget = QtWidgets.QListWidget(self.groupBox_3)
        self.listWidget.setGeometry(QtCore.QRect(20, 430, 381, 51))
        self.listWidget.setStyleSheet("background-color: rgb(217, 217, 217);\n"
                                      "color: rgb(0, 0, 0);\n"
                                      "border-radius: 10px; ")
        self.listWidget.setObjectName("listWidget")
        self.del_select = QtWidgets.QPushButton(self.groupBox_3)
        self.del_select.setGeometry(QtCore.QRect(220, 490, 181, 51))
        self.del_select.setStyleSheet("")
        self.del_select.setCheckable(False)
        self.del_select.setObjectName("del_select")
        self.sending = QtWidgets.QPushButton(self.groupBox_3)
        self.sending.setGeometry(QtCore.QRect(120, 550, 181, 51))
        self.sending.setStyleSheet("")
        self.sending.setCheckable(False)
        self.sending.setObjectName("sending")
        self.pushButton_4 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_4.setGeometry(QtCore.QRect(840, 840, 181, 51))
        self.pushButton_4.setStyleSheet("")
        self.pushButton_4.setCheckable(False)
        self.pushButton_4.setObjectName("pushButton_4")
        MainForm.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainForm)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1151, 26))
        self.menubar.setObjectName("menubar")
        self.menu = QtWidgets.QMenu(self.menubar)
        self.menu.setObjectName("menu")
        self.helping = QtWidgets.QMenu(self.menubar)
        self.helping.setObjectName("helping")
        self.menu_3 = QtWidgets.QMenu(self.menubar)
        self.menu_3.setObjectName("menu_3")
        MainForm.setMenuBar(self.menubar)
        self.load = QtWidgets.QAction(MainForm)
        self.load.setObjectName("load")
        self.save = QtWidgets.QAction(MainForm)
        self.save.setObjectName("save")
        self.actionQuit = QtWidgets.QAction(MainForm)
        self.actionQuit.setObjectName("actionQuit")
        self.user_inf = QtWidgets.QAction(MainForm)
        self.user_inf.setObjectName("user_inf")
        self.prog_inf = QtWidgets.QAction(MainForm)
        self.prog_inf.setObjectName("prog_inf")
        self.menu.addAction(self.load)
        self.menu.addSeparator()
        self.menu.addAction(self.save)
        self.menu.addSeparator()
        self.menu.addAction(self.actionQuit)
        self.menu_3.addAction(self.user_inf)
        self.menu_3.addAction(self.prog_inf)
        self.menubar.addAction(self.menu.menuAction())
        self.menubar.addAction(self.menu_3.menuAction())
        self.load.setShortcut("Ctrl+O")
        self.save.setShortcut("Ctrl+S")
        self.user_inf.setShortcut("Ctrl+U")
        self.prog_inf.setShortcut("Ctrl+P")

        self.retranslateUi(MainForm)
        QtCore.QMetaObject.connectSlotsByName(MainForm)

        self.table_design()
        # Добавляем кнопкам функции
        self.add.clicked.connect(self.add_row)
        self.sending.clicked.connect(self.send_email)
        self.attach.clicked.connect(self.attach_file)
        self.del_select.clicked.connect(self.remove_file)
        self.del_id.clicked.connect(self.remove_row)
        self.save.triggered.connect(self.save_file)
        self.load.triggered.connect(self.open_file)
        self.pushButton_4.clicked.connect(QtWidgets.qApp.quit)
        self.del_text.clicked.connect(self.entry_inf)
        self.actionQuit.triggered.connect(QtWidgets.QApplication.quit)


    def remove_file(self): # Удаление файла из прикрепленных
        for item in self.listWidget.selectedItems():
            file_path = self.file_paths[self.listWidget.row(item)]
            self.file_paths.remove(file_path)
            self.listWidget.takeItem(self.listWidget.row(item))


    def check_internet_connection(self): #Проверка соединения к интернету
        try:
            # Попытка установить соединение с хорошо известным IP-адресом (например, адрес Google)
            socket.create_connection(("www.google.com", 80), timeout=3)
            return True
        except OSError:
            pass
        return False

    def win_err(self, head, text, iconik, name_button,second_button=False):
        # Функция для создания определенного сообщения
        message_box = QMessageBox(QMessageBox.Warning, head, text)
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setWindowIcon(QIcon("resources/mail_ikonka2.png"))
        message_box.setIcon(iconik)
        message_box.button(QMessageBox.Ok).setText(name_button)

        if second_button:
            cancel_button = message_box.addButton(QMessageBox.Cancel)
            cancel_button.setText("Отмена")

        message_box.exec_()

        if second_button and message_box.clickedButton() == cancel_button:
            return "Отмена"

    def attach_file(self): # Добавление файла к прикрепленным
        file_path, _ = QFileDialog.getOpenFileName(self, 'Выберите файл')
        if file_path:
            self.file_paths.append(file_path)
            item = QListWidgetItem(QIcon(file_path), os.path.basename(file_path))
            self.listWidget.addItem(item)

    def open_file(self): # Открытие таблицы excel
        file_dialog = QFileDialog(self)
        filepath, _ = file_dialog.getOpenFileName(self, "Открыть файл Excel", "", "Excel Files (*.xlsx *.xls)")
        if filepath:
            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active

                self.tableWidget.setRowCount(sheet.max_row)
                self.tableWidget.setColumnCount(min(sheet.max_column, 3))

                for row in range(1, sheet.max_row + 1):
                    for col in range(1, min(sheet.max_column + 1, 4)):
                        cell_value = sheet.cell(row=row, column=col).value
                        item = QTableWidgetItem(str(cell_value))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.tableWidget.setItem(row - 1, col - 1, item)
            except Exception:
                self.win_err("Ошибка", f"Ошибка при открытии файла", QMessageBox.Warning, "Хорошо")
                return

    def entry_inf(self): # Удаление входных данных
        self.lineEdit_2.clear()
        self.lineEdit_3.clear()
        self.lineEdit_4.clear()
        self.lineEdit_5.clear()
        self.lineEdit_6.clear()


    def save_file(self): # Сохранение таблицы Excel
        file_dialog = QFileDialog(self)
        filepath, _ = file_dialog.getSaveFileName(self, "Сохранить файл Excel", "", "Excel Files (*.xlsx *.xls)")
        if filepath:
            workbook = Workbook()
            sheet = workbook.active

            for row in range(self.tableWidget.rowCount()):
                for col in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(row, col)
                    if item is not None:
                        cell_value = item.text()
                        sheet.cell(row=row + 1, column=col + 1).value = cell_value

            # Установка ширины столбцов
            for col in range(self.tableWidget.columnCount()):
                column_letter = get_column_letter(col + 1)
                column_width = self.tableWidget.columnWidth(col)
                sheet.column_dimensions[column_letter].width = column_width / 6.0  # Преобразование ширины в пикселях в ширину столбца в Excel

            workbook.save(filepath)

    def table_design(self): # Настройка таблицы в программе
        # Установка количества строк и колонок
        self.tableWidget.setRowCount(0)  # Пока строк нету
        self.tableWidget.setColumnCount(3)  # 3 колонки

        # Задание размеров колонок
        self.tableWidget.setColumnWidth(0, 200)  # Колонка 1 - Фамилия Инициалы
        self.tableWidget.setColumnWidth(1, 150)  # Колонка 2 - Группа студента
        self.tableWidget.setColumnWidth(2, 290)  # Колонка 3 - Email 280-298(если нет строк)

        # Установка обводки ячеек и черного текста
        self.tableWidget.setStyleSheet("background-color: rgb(217, 217, 217);"
                                       "color: black;")
        self.tableWidget.horizontalHeader().setStyleSheet("border-color: black; color: black;")
        self.tableWidget.horizontalHeader().setSectionResizeMode(
            QHeaderView.Fixed)  # Неизменяемость верхних заголовков
        self.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.tableWidget.verticalHeader().setStyleSheet("border-color: black; color: black;")
        # Установка названий колонок
        self.tableWidget.setHorizontalHeaderLabels(["ФИО", "Группа студента", "Email"])



    def add_row(self): # Добавление информации о студенте в таблицу
        row_position = self.tableWidget.rowCount()
        lineEdits = [self.lineEdit_2, self.lineEdit_3, self.lineEdit_4, self.lineEdit_5, self.lineEdit_6]
        if any(not lineEdit.text() for lineEdit in lineEdits):
            self.win_err("Ошибка", "Заполните все поля для добавления", QMessageBox.Warning, "Хорошо")
            return
        if not self.lineEdit_2.text().isalpha() or not self.lineEdit_3.text().isalpha() or not self.lineEdit_4.text().isalpha():
            self.win_err("Ошибка", "Заполните ФИО правильно ", QMessageBox.Warning, "Хорошо")
            return
        elif not re.match(r'^\d{8}$', self.lineEdit_5.text()):
            self.win_err("Ошибка", "Заполните группу правильно (8 цифр) ", QMessageBox.Warning, "Хорошо")
            return
        elif not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', self.lineEdit_6.text()):
            self.win_err("Ошибка", "Заполните email правильно ", QMessageBox.Warning, "Хорошо")
            return

        self.tableWidget.insertRow(row_position)
        item1_text = self.lineEdit_2.text()
        item2_text = self.lineEdit_3.text()
        item3_text = self.lineEdit_4.text()
        item4_text = self.lineEdit_5.text()
        item5_text = self.lineEdit_6.text()

        item1 = QtWidgets.QTableWidgetItem(f"{item1_text} {item2_text[0]}.{item3_text[0]}.")
        item1.setTextAlignment(Qt.AlignCenter)
        item2 = QtWidgets.QTableWidgetItem(f"{item4_text}")
        item2.setTextAlignment(Qt.AlignCenter)
        item3 = QtWidgets.QTableWidgetItem(f"{item5_text}")
        item3.setTextAlignment(Qt.AlignCenter)

        # Вставка элементов в таблицу
        self.tableWidget.setItem(row_position, 0, item1)
        self.tableWidget.setItem(row_position, 1, item2)
        self.tableWidget.setItem(row_position, 2, item3)



    def remove_row(self): # Удаление информации о студенте из таблицы
        row_number_text = self.lineEdit.text()  # Получаем номер строки из QLineEdit
        if not row_number_text:  # Проверяем, что строка не пустая
            self.win_err("Предупреждение", "Не указан номер строки.", QMessageBox.Warning, "Хорошо")
            return
        try:
            row_number = int(row_number_text) - 1  # Преобразуем номер строки в целое число
            row_count = self.tableWidget.rowCount()  # Получаем общее количество строк в таблице
            if row_count == 0:  # Проверяем, что таблица пуста
                self.win_err("Предупреждение", "Таблица не заполнена", QMessageBox.Warning, "Хорошо")
            elif 0 <= row_number < row_count:  # Проверяем, что номер строки в допустимом диапазоне
                result = self.win_err("Препреждение", "Вы уверенны, что хотите удалить\nданного студента из таблицы?", QMessageBox.Warning, "Конечно",second_button=True)
                if result == "Отмена":
                    return
                self.tableWidget.removeRow(row_number)  # Удаляем строку с указанным номером
                self.win_err("Успех", "Строка успешно удалена.", QMessageBox.Information, "Хорошо")
            else:
                self.win_err("Ошибка", "Недопустимый номер строки.", QMessageBox.Critical, "Хорошо")
        except ValueError:
            self.win_err("Ошибка", "Неверный формат номера строки.", QMessageBox.Critical, "Хорошо")

    def send_email(self): # Отправка Email студентам, зарегистрированным в таблице
        if self.check_internet_connection():
            """
            Следующие две строки необходимо было заполнить 
            пользователю. Первая строка - это пароль.
            Вторая строка - это почта, через которую пользователь
            будет отправлять email
            """
            sender_password = 'rsjv dwhl ystk tjnz'
            sender_email = 'gogle35673@gmail.com'
            subject = self.lineEdit_7.text()
            message = self.plainTextEdit.toPlainText()
            if self.tableWidget.rowCount() == 0:
                self.win_err("Предупреждение", "Таблица не заполнена", QMessageBox.Warning, "Хорошо")
                return
            try:
                server = smtplib.SMTP("smtp.gmail.com", 587)
                server.starttls()
                server.login(sender_email, sender_password)
                recipient = self.lineEdit_9.text()
                selected_rows = self.tableWidget.selectionModel().selectedRows()
                if len(selected_rows) > 0:
                    # Отправить email выбранным пользователям из выделенных строк
                    for row in selected_rows:
                        receiver_email = self.tableWidget.item(row.row(), 2).text()

                        msg = MIMEMultipart()
                        msg['From'] = sender_email
                        msg['To'] = receiver_email
                        msg['Subject'] = subject

                        body = MIMEText(message, 'plain', 'utf-8')
                        msg.attach(body)

                        for file_path in self.file_paths:
                            attachment = open(file_path, 'rb')
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload((attachment).read())
                            encoders.encode_base64(part)
                            part.add_header('Content-Disposition',
                                            "attachment; filename= %s" % os.path.basename(file_path))
                            msg.attach(part)

                        server.send_message(msg)
                elif recipient == "Всем":
                    # Отправить email всем пользователям таблицы
                    for row in range(self.tableWidget.rowCount()):
                        receiver_email = self.tableWidget.item(row, 2).text()

                        msg = MIMEMultipart()
                        msg['From'] = sender_email
                        msg['To'] = receiver_email
                        msg['Subject'] = subject

                        body = MIMEText(message, 'plain', 'utf-8')
                        msg.attach(body)

                        for file_path in self.file_paths:
                            attachment = open(file_path, 'rb')
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload((attachment).read())
                            encoders.encode_base64(part)
                            part.add_header('Content-Disposition',
                                            "attachment; filename= %s" % os.path.basename(file_path))
                            msg.attach(part)

                        server.send_message(msg)
                else:
                    # Искать группу во втором столбце и отправлять email участникам этой группы
                    group_found = False
                    for row in range(self.tableWidget.rowCount()):
                        group = self.tableWidget.item(row, 1).text()
                        if group == recipient:
                            group_found = True
                            receiver_email = self.tableWidget.item(row, 2).text()

                            msg = MIMEMultipart()
                            msg['From'] = sender_email
                            msg['To'] = receiver_email
                            msg['Subject'] = subject

                            body = MIMEText(message, 'plain', 'utf-8')
                            msg.attach(body)

                            for file_path in self.file_paths:
                                attachment = open(file_path, 'rb')
                                part = MIMEBase('application', 'octet-stream')
                                part.set_payload((attachment).read())
                                encoders.encode_base64(part)
                                part.add_header('Content-Disposition',
                                                "attachment; filename= %s" % os.path.basename(file_path))
                                msg.attach(part)

                            server.send_message(msg)

                    if not group_found:
                        self.win_err("Ошибка", "Нет такой группы в таблице.", QMessageBox.Warning, "Хорошо")
                        return

                server.quit()
                self.win_err("Успех", "Сообщение отправлено успешно.", QMessageBox.Information, "Хорошо")
            except Exception as e:
                self.win_err("Ошибка", "Произошла ошибка при отправке сообщения.", QMessageBox.Warning, "Хорошо")
                return

        else:
            self.win_err("Ошибка", "Отсутствует подключение к интернету.\nПроверьте соединение", QMessageBox.Warning, "Хорошо")

    def retranslateUi(self, MainForm): # Дизайн окна
        _translate = QtCore.QCoreApplication.translate
        MainForm.setWindowTitle(_translate("MainForm", "Информирование студентов"))
        self.groupBox.setTitle(_translate("MainForm", "Данные студентов"))
        self.label_2.setText(_translate("MainForm", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt;\">Фамилия</span></p></body></html>"))
        self.label_3.setText(_translate("MainForm", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt;\">Имя</span></p></body></html>"))
        self.label_4.setText(_translate("MainForm", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt;\">Отчество</span></p></body></html>"))
        self.label_5.setText(_translate("MainForm", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt;\">Группа</span></p></body></html>"))
        self.label_6.setText(_translate("MainForm", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt;\">Email</span></p></body></html>"))
        self.add.setText(_translate("MainForm", "Добавить"))
        self.del_text.setText(_translate("MainForm", "Стереть"))
        self.groupBox_2.setTitle(_translate("MainForm", "Удаление данных студента"))
        self.label.setText(_translate("MainForm", "<html><head/><body><p align=\"center\"><span style=\" font-size:11pt;\">Введите номер строки студента</span></p></body></html>"))
        self.del_id.setText(_translate("MainForm", "Удалить"))
        self.groupBox_3.setTitle(_translate("MainForm", "Рассылка сообщений студентам"))
        self.label_7.setText(_translate("MainForm", "<html><head/><body><p align=\"center\"><span style=\" font-size:9pt; font-weight:600;\">Введите тему сообщения(необязательно)</span></p><p align=\"center\"><span style=\" font-weight:600;\"><br/></span></p></body></html>"))
        self.sending.setText(_translate("MainForm", "Отправить"))
        self.label_8.setText(_translate("MainForm", "<html><head/><body><p align=\"center\"><span style=\" font-size:9pt; font-weight:600;\">Введите сообщение</span></p><p align=\"center\"><br/></p></body></html>"))
        self.label_9.setText(_translate("MainForm", "<html><head/><body><p><span style=\" font-weight:600;\">Введите номер группы, &quot;Всем&quot; или выберите студентов</span></p></body></html>"))
        self.attach.setText(_translate("MainForm", "Прикрепить"))
        self.del_select.setText(_translate("MainForm", "Открепить"))
        self.pushButton_4.setText(_translate("MainForm", "Выход"))
        self.menu.setTitle(_translate("MainForm", "Файл"))
        self.menu_3.setTitle(_translate("MainForm", "Информация"))
        self.load.setText(_translate("MainForm", "Открыть"))
        self.load.setIconText(_translate("MainForm", "Открыть            "))
        self.save.setText(_translate("MainForm", "Сохранить"))
        self.actionQuit.setText(_translate("MainForm", "Выйти"))
        self.user_inf.setText(_translate("MainForm", "Об авторе"))
        self.prog_inf.setText(_translate("MainForm", "О программе"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainForm = QtWidgets.QMainWindow()
    ui = Ui_MainForm()
    ui.setupUi(MainForm)
    MainForm.show()
    sys.exit(app.exec_())
